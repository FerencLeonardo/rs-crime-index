import re
import os
import pandas as pd

from openpyxl import load_workbook

# Main function
def main():
    directory_path = "/home/leoli/projects/rs-crime-index/data/excel"
    files_list = get_all_files_list(directory_path)

    directory_path = "/home/leoli/projects/rs-crime-index/data"

    create_csv(directory_path, files_list)
    return 0

# <----- Single-use functions ----->
def _rename_all_files(directory_path, files_list):
    for file in files_list:
        filename = file[len(directory_path):]
        match = re.search(r"\d{4}", filename)
        if match:
            year = match.group()
            os.rename(file, os.path.join("data/", year + ".xlsx"))
        else:
            raise ValueError(f"No 4-digit year found in filename: {file}")

# This function is very slow!
def _remove_sheet(files_list, sheet):
    for file in files_list:
        wb = load_workbook(file, read_only = True)
        wb.close()
        wb = load_workbook(file)
        if sheet in wb.sheetnames:
            del wb[sheet]
            wb.save(file)
        wb.close()

# <----- Helper functions ----->
def _get_year(file):
    match = re.search(r"\d{4}", file)
    if match:
        year = match.group()
        return year
    else:
        raise ValueError(f"No 4-digit year found in filename: {file}")

def _get_file(files_list, year):
    # Loop over files_list and find the first file that contains the year.
    for file in files_list:
        if _get_year(file) == year:
            return file
    raise ValueError(f"{year} is not in the files_list")

def _get_row_of_word(file, sheet, word):
    row = None
    df = pd.read_excel(
        file,
        sheet_name = sheet,
        header = None
    )
    old_first_name = df.columns[0]
    df = df.rename(columns={old_first_name: "values"})
    for i, val in enumerate(df["values"]):
        if val == word:
            row = i
            break
    return row

# <----- Primary functions ----->
def get_all_files_list(directory_path):
    files_list = []
    for _, _, filenames in os.walk(directory_path):
        for filename in filenames:
            files_list.append(os.path.join(directory_path, filename))
    files_list.sort()
    return files_list

def create_clean_df():
    clean_df = pd.DataFrame(columns=["city", "category", "year", "month", "count"])
    return clean_df

def get_sheet_df(file, sheet):
    df = pd.DataFrame()

    # Open file with pandas to read data
    excel_file = pd.ExcelFile(file)
    if sheet in excel_file.sheet_names:
        head_row = _get_row_of_word(file, sheet, "Municípios")
        tail_row = _get_row_of_word(file, sheet, "Total RS")

        df = pd.read_excel(
            file,
            sheet_name = sheet,
            usecols="A:O",
            header = head_row,
            nrows = tail_row - head_row - 1,
        )
    return df

def add_to_clean_df(clean_df, sheet_df, year, month):
    if sheet_df.empty:
        return clean_df

    df = sheet_df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    city_col = df.columns[0]
    long = df.melt(
        id_vars=[city_col],
        var_name="category",
        value_name="count",
    ).rename(columns={city_col: "city"}) # Rename the city column to city

    long["city"] = long["city"].astype("string").str.strip()
    long["category"] = (
        long["category"].astype("string").str.replace(r"\s+", " ", regex=True).str.strip()
    )

    long["year"] = int(year) # Add year column with the year value
    long["month"] = int(month) # Add month column with the month value

    long = long[["city", "category", "year", "month", "count"]] # Re-organize the columns in correct order

    if clean_df.empty:
        return long
    return pd.concat([clean_df, long], ignore_index=True)

def create_csv(directory_path, files_list):
    months = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]

    csv_path = os.path.join(directory_path, "csv")
    os.makedirs(csv_path, exist_ok=True)

    clean_df = create_clean_df()

    for file in files_list:
        for month in months:
            sheet_df = get_sheet_df(file, month)
            if not sheet_df.empty:
                year = int(_get_year(file))
                month_num = months.index(month) + 1
                clean_df = add_to_clean_df(clean_df, sheet_df, year, month_num)
    
    clean_df = clean_df.sort_values(by=["city", "year", "month"])
    clean_df.to_csv(
        os.path.join(csv_path, "rs-crime-index.csv"),
        index = False,
        encoding = "utf-8",
    )
                
if __name__ == "__main__":
    main()


